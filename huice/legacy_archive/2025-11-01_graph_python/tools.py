import asyncio
import os
from typing import List, Dict, Any
from datetime import datetime

from langchain_core.tools import tool
from langchain_mcp_adapters.client import MultiServerMCPClient
# pip install openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

@tool
def get_weather(city: str) -> str:
    """Get weather for a given city."""
    return f"{city}，今天是晴天，温度为 25 摄氏度。!"

def get_zhipu_search_mcp_tools():
    client = MultiServerMCPClient(
        {
            "search": {
                "url": "https://open.bigmodel.cn/api/mcp/web_search/sse?Authorization=0d3c7b3d55c84d6682663dbdbdb3d614.cpoJIoZ3NFaqXARQ",
                "transport": "sse",
            }
        }
    )
    tools = asyncio.run(client.get_tools())
    return tools

def get_tavily_search_tools():
    client = MultiServerMCPClient(
        {
            "search": {
                "url": "https://mcp.tavily.com/mcp/?tavilyApiKey=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",
                "transport": "streamable_http",
            }
        }
    )
    tools = asyncio.run(client.get_tools())
    return tools

def get_playwright_mcp_tools():
    client = MultiServerMCPClient(
        {
            "playwright_mcp": {
                "command": "npx",
                "args": ["@playwright/mcp@latest"],
                "transport": "stdio",
            }
        }
    )
    tools = asyncio.run(client.get_tools())
    return tools

def get_chrome_devtools_mcp_tools():
    client = MultiServerMCPClient(
        {
            "chrome_devtools_mcp": {
                "command": "npx",
                "args": ["chrome-devtools-mcp@latest", "--headless=false", "--isolated=true"],
                "transport": "stdio",
            }
        }
    )
    tools = asyncio.run(client.get_tools())
    return tools


def get_chrome_mcp_tools():
    client = MultiServerMCPClient(
        {
            "chrome_mcp": {
                "url": "http://127.0.0.1:12306/mcp",
                "transport": "streamable_http",
            }
        }
    )
    tools = asyncio.run(client.get_tools())
    return tools

def get_mcp_server_chart_tools():
    client = MultiServerMCPClient(
        {
            "mcp_chart_server": {
                "command": "npx",
                "args": ["-y", "@antv/mcp-server-chart"],
                "transport": "stdio",
            }
        }
    )
    tools = asyncio.run(client.get_tools())
    return tools
# print(get_chrome_devtools_mcp_tools())
# tools = get_zhipu_search_mcp_tools()
# print(tools)


# https://app.tavily.com/home
os.environ["TAVILY_API_KEY"] = "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"
# # pin install langchain-tavily
from langchain_tavily import TavilySearch
toolSearch = TavilySearch(max_results=2)


@tool
def save_test_cases_to_excel(
    test_cases: List[Dict[str, Any]],
    file_path: str = "test_cases.xlsx",
    sheet_name: str = "测试用例",
    append: bool = False,
    columns: List[str] = None
) -> str:
    """
    将测试用例保存到Excel文件，支持专业格式化。

    此函数可以创建新的Excel文件或向现有文件追加测试用例数据，
    具有自动格式化、列宽调整和完善的错误处理功能。

    参数:
        test_cases: 测试用例列表，每个测试用例是一个字典。
                   示例结构:
                   [
                       {
                           "用例ID": "TC001",
                           "用例标题": "登录功能测试",
                           "前置条件": "用户已注册",
                           "测试步骤": "1. 打开登录页面\n2. 输入用户名和密码\n3. 点击登录按钮",
                           "预期结果": "成功登录并跳转到首页",
                           "优先级": "高",
                           "用例类型": "功能测试"
                       },
                       ...
                   ]
        file_path: Excel文件保存路径，默认为 "test_cases.xlsx"
        sheet_name: 工作表名称，默认为 "测试用例"
        append: 是否追加到现有文件，True表示追加，False表示创建新文件。默认为False
        columns: 可选的列名列表，用于指定列的顺序。
                如果为None，则自动从测试用例的键中检测列名。
                示例: ["用例ID", "用例标题", "优先级", "预期结果"]

    返回:
        str: 成功消息（包含文件路径和保存的测试用例数量）或错误消息

    异常:
        不会抛出异常，所有错误都会被捕获并作为错误消息返回

    使用示例:
        >>> test_data = [
        ...     {"用例ID": "TC001", "用例标题": "登录测试", "优先级": "高"},
        ...     {"用例ID": "TC002", "用例标题": "登出测试", "优先级": "中"}
        ... ]
        >>> result = save_test_cases_to_excel(test_data, "我的测试.xlsx")
        >>> print(result)
        成功创建！已保存 2 条测试用例到文件: /path/to/我的测试.xlsx
    """
    try:
        # 验证输入数据
        if not test_cases:
            return "错误：测试用例列表为空"

        if not isinstance(test_cases, list):
            return "错误：test_cases 必须是字典列表"

        # 确定要使用的列
        if columns is None:
            # 从所有测试用例中自动检测列名
            all_keys = []
            for case in test_cases:
                if not isinstance(case, dict):
                    return f"错误：每个测试用例必须是字典类型，当前发现 {type(case)}"
                for key in case.keys():
                    if key not in all_keys:
                        all_keys.append(key)
            columns = all_keys

        if not columns:
            return "错误：在测试用例中未找到任何列"

        # 定义Excel样式
        # 表头样式：白色粗体文字，蓝色背景
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 单元格样式：普通字体，左对齐，自动换行
        cell_font = Font(name='Arial', size=10)
        cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        # 边框样式：细黑色边框
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # 判断是追加模式还是创建新文件
        if append and os.path.exists(file_path):
            # 追加模式：加载现有文件
            try:
                wb = load_workbook(file_path)
                if sheet_name in wb.sheetnames:
                    # 工作表已存在，追加到末尾
                    ws = wb[sheet_name]
                    start_row = ws.max_row + 1
                else:
                    # 工作表不存在，创建新工作表
                    ws = wb.create_sheet(sheet_name)
                    start_row = 1
            except Exception as e:
                return f"错误：加载现有文件失败: {str(e)}"
        else:
            # 创建新文件
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            start_row = 1

        # 如果是新工作表，写入表头行
        if start_row == 1:
            for col_idx, header in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            start_row = 2  # 数据从第2行开始

        # 写入测试用例数据
        for row_idx, test_case in enumerate(test_cases, start=start_row):
            for col_idx, column_name in enumerate(columns, start=1):
                # 获取字典中的值，如果不存在则为空字符串
                value = test_case.get(column_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = thin_border

        # 根据内容类型自动调整列宽
        for col_idx, column_name in enumerate(columns, start=1):
            column_letter = ws.cell(row=1, column=col_idx).column_letter

            # 根据列名模式设置默认宽度
            # ID、优先级、类型、状态等短字段：15个字符宽度
            if any(keyword in column_name.lower() for keyword in ['id', 'priority', 'type', 'status', '优先级', '类型', '状态']):
                ws.column_dimensions[column_letter].width = 15
            # 标题、名称、摘要等中等字段：30个字符宽度
            elif any(keyword in column_name.lower() for keyword in ['title', 'name', 'summary', '标题', '名称', '摘要']):
                ws.column_dimensions[column_letter].width = 30
            # 步骤、结果、描述、详情等长字段：50个字符宽度
            elif any(keyword in column_name.lower() for keyword in ['step', 'result', 'description', 'detail', '步骤', '结果', '描述', '详情', '条件']):
                ws.column_dimensions[column_letter].width = 50
            # 其他字段：25个字符宽度
            else:
                ws.column_dimensions[column_letter].width = 25

        # 设置行高以提高可读性
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 30

        # 保存文件
        try:
            wb.save(file_path)
        except PermissionError:
            return f"错误：权限被拒绝。文件 '{file_path}' 可能正在其他程序中打开。"
        except Exception as e:
            return f"错误：保存文件失败: {str(e)}"

        # 生成成功消息
        mode_text = "追加" if (append and os.path.exists(file_path)) else "创建"
        abs_path = os.path.abspath(file_path)
        return f"成功{mode_text}！已保存 {len(test_cases)} 条测试用例到文件: {abs_path}"

    except Exception as e:
        return f"未预期的错误: {str(e)}"